from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                               QLabel, QTableWidget, QTableWidgetItem, QFrame,
                               QComboBox, QFileDialog, QMessageBox, QScrollArea)
from PySide6.QtCore import Qt, QMargins
from PySide6.QtCharts import QChart, QChartView, QPieSeries
from PySide6.QtGui import QPainter, QColor
from desktop.controllers.reports_controller import ReportsController
from datetime import date

_MONTHS_PT = ['', 'Janeiro', 'Fevereiro', 'Marco', 'Abril', 'Maio', 'Junho',
              'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

_BTN_ACTIVE = (
    "QPushButton { background-color: #2c3e50; color: white; border: none;"
    " border-radius: 4px; padding: 4px 14px; font-size: 12px; font-weight: bold; }"
)
_BTN_INACTIVE = (
    "QPushButton { background-color: #dfe6e9; color: #636e72; border: none;"
    " border-radius: 4px; padding: 4px 14px; font-size: 12px; }"
    "QPushButton:hover { background-color: #b2bec3; }"
)

_COLORS = [
    "#3498db", "#e74c3c", "#2ecc71", "#f39c12", "#9b59b6",
    "#1abc9c", "#e67e22", "#e91e63", "#00bcd4", "#8bc34a",
    "#ff5722", "#607d8b", "#795548", "#ff9800", "#673ab7",
]


class ReportsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.controller = ReportsController()
        self.current_date = date.today()
        self._monthly_data = {'top_products': [], 'top_customers': []}
        self._products_sort = 'revenue'
        self._customers_sort = 'total_spent'
        self.setup_ui()

    def setup_ui(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        self.setStyleSheet("background-color: #ecf0f1;")

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { background-color: #ecf0f1; border: none; }")

        scroll_widget = QWidget()
        scroll_widget.setStyleSheet("background-color: #ecf0f1;")
        layout = QVBoxLayout(scroll_widget)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(18)

        scroll.setWidget(scroll_widget)
        root_layout.addWidget(scroll)

        # ── Header ──
        header_layout = QHBoxLayout()

        title = QLabel("Relatorio do Dia")
        title.setStyleSheet("font-size: 32px; font-weight: bold; color: #2c3e50; background-color: transparent;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        btn_excel = QPushButton("Exportar Excel")
        btn_excel.setFixedHeight(42)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #27ae60; color: white;
                font-size: 15px; padding: 0 20px;
                border-radius: 5px; font-weight: bold;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        btn_excel.clicked.connect(self.export_excel)
        header_layout.addWidget(btn_excel)

        layout.addLayout(header_layout)

        # ── Seletor de data ──
        date_layout = QHBoxLayout()

        date_label = QLabel("Data:")
        date_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #2c3e50;")
        date_layout.addWidget(date_label)

        self.date_combo = QComboBox()
        self.date_combo.setFixedHeight(36)
        self.date_combo.setStyleSheet("""
            QComboBox {
                padding: 0 12px;
                font-size: 14px;
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
                color: #2c3e50;
                min-width: 200px;
            }
        """)
        self.date_combo.currentIndexChanged.connect(self.on_date_changed)
        date_layout.addWidget(self.date_combo)

        date_layout.addStretch()
        layout.addLayout(date_layout)

        # ── Cards de resumo ──
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(15)

        self.card_orders  = self._create_summary_card("Comandas Fechadas", "0",        "#3498db")
        self.card_revenue = self._create_summary_card("Receita Total",     "R$ 0,00",  "#27ae60")
        self.card_items   = self._create_summary_card("Itens Vendidos",    "0",        "#9b59b6")

        cards_layout.addWidget(self.card_orders)
        cards_layout.addWidget(self.card_revenue)
        cards_layout.addWidget(self.card_items)

        layout.addLayout(cards_layout)

        # ── Secao: Vendas por Categoria ──
        layout.addWidget(self._create_section(
            "Vendas por Categoria",
            ["Categoria", "Quantidade", "Receita"],
            [160, 120, 120],
            "category_table", "category_chart_view",
        ))

        # ── Secao: Vendas por Produto ──
        layout.addWidget(self._create_section(
            "Vendas por Produto",
            ["Produto", "Categoria", "Quantidade", "Receita"],
            [240, 140, 110, 120],
            "product_table", "product_chart_view",
        ))

        # ── Secao: Comandas Fechadas ──
        layout.addWidget(self._create_section(
            "Comandas Fechadas",
            ["ID", "Cliente", "Horario", "Total"],
            [60, 220, 120, 120],
            "orders_table", "orders_chart_view",
        ))

        # ── Divisor visual ──
        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.HLine)
        divider.setStyleSheet("color: #bdc3c7;")
        layout.addWidget(divider)

        # ── Secao: TOP do Mes ──
        top_frame = QFrame()
        top_frame.setStyleSheet("background-color: white; border-radius: 10px;")
        top_layout = QVBoxLayout(top_frame)
        top_layout.setContentsMargins(18, 14, 18, 18)
        top_layout.setSpacing(12)

        # Cabecalho do TOP com seletor de mes
        top_header = QHBoxLayout()
        top_title = QLabel("TOP do Mes")
        top_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        top_header.addWidget(top_title)
        top_header.addStretch()

        month_label = QLabel("Mes:")
        month_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        top_header.addWidget(month_label)

        self.month_combo = QComboBox()
        self.month_combo.setFixedHeight(34)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 0 12px;
                font-size: 14px;
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
                color: #2c3e50;
                min-width: 180px;
            }
        """)
        self.month_combo.currentIndexChanged.connect(self._on_month_changed)
        top_header.addWidget(self.month_combo)
        top_layout.addLayout(top_header)

        # Dois paineis lado a lado: Top Produtos | Top Clientes
        panels_layout = QHBoxLayout()
        panels_layout.setSpacing(16)
        panels_layout.addWidget(self._create_top_panel(
            "Top 5 Produtos mais vendidos",
            ["Produto", "Qtd", "Receita"],
            [200, 60, 90],
            "top_products_table", "top_products_chart_view",
            ["Quantidade", "Receita"], "prod",
        ))
        panels_layout.addWidget(self._create_top_panel(
            "Top 5 Clientes que mais gastaram",
            ["Cliente", "Comandas", "Total"],
            [180, 80, 90],
            "top_customers_table", "top_customers_chart_view",
            ["Comandas", "Total Gasto"], "cust",
        ))

        self._prod_btn_a.clicked.connect(lambda: self._set_sort('products', 'quantity'))
        self._prod_btn_b.clicked.connect(lambda: self._set_sort('products', 'revenue'))
        self._cust_btn_a.clicked.connect(lambda: self._set_sort('customers', 'total_orders'))
        self._cust_btn_b.clicked.connect(lambda: self._set_sort('customers', 'total_spent'))
        top_layout.addLayout(panels_layout)
        layout.addWidget(top_frame)

    # ── HELPERS ──────────────────────────────────────────────────────

    def _create_summary_card(self, title, value, color):
        card = QFrame()
        card.setFixedHeight(90)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 10px;
            }}
        """)
        card_layout = QHBoxLayout(card)
        card_layout.setContentsMargins(20, 0, 20, 0)
        card_layout.setSpacing(20)

        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)

        title_label = QLabel(title)
        title_label.setStyleSheet("color: rgba(255,255,255,0.85); font-size: 13px; background-color: transparent;")
        text_layout.addWidget(title_label)

        value_label = QLabel(value)
        value_label.setStyleSheet("color: white; font-size: 28px; font-weight: bold; background-color: transparent;")
        value_label.setObjectName("value_label")
        text_layout.addWidget(value_label)

        card_layout.addLayout(text_layout)
        card_layout.addStretch()

        return card

    def _make_chart_view(self):
        chart = QChart()
        chart.setAnimationOptions(QChart.AnimationOption.SeriesAnimations)
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignmentFlag.AlignBottom)
        chart.setBackgroundVisible(False)
        chart.setMargins(QMargins(4, 4, 4, 4))

        view = QChartView(chart)
        view.setRenderHint(QPainter.RenderHint.Antialiasing)
        view.setStyleSheet("background: transparent; border: none;")
        view.setMinimumHeight(280)
        return view

    def _create_section(self, section_title, headers, col_widths, table_attr, chart_attr):
        frame = QFrame()
        frame.setStyleSheet("background-color: white; border-radius: 10px;")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setSpacing(10)
        frame_layout.setContentsMargins(18, 14, 18, 14)

        title_label = QLabel(section_title)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        frame_layout.addWidget(title_label)

        # Tabela (esquerda) + Grafico (direita)
        h_layout = QHBoxLayout()
        h_layout.setSpacing(16)

        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                font-size: 14px;
                color: #2c3e50;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
                font-size: 13px;
            }
            QTableWidget::item {
                color: #2c3e50;
                padding: 6px 8px;
            }
        """)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(38)
        table.horizontalHeader().setStretchLastSection(True)

        for i, w in enumerate(col_widths):
            table.setColumnWidth(i, w)

        h_layout.addWidget(table, stretch=55)

        chart_view = self._make_chart_view()
        h_layout.addWidget(chart_view, stretch=45)

        frame_layout.addLayout(h_layout)

        setattr(self, table_attr, table)
        setattr(self, chart_attr, chart_view)

        return frame

    def _create_top_panel(self, section_title, headers, col_widths,
                          table_attr, chart_attr, toggle_labels=None, toggle_prefix=None):
        """Painel vertical: titulo + botoes toggle + grafico + tabela."""
        frame = QFrame()
        frame.setStyleSheet("background-color: #f8f9fa; border-radius: 8px;")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(14, 12, 14, 12)
        frame_layout.setSpacing(8)

        # Cabecalho: titulo + botoes de filtro
        header_row = QHBoxLayout()
        header_row.setSpacing(6)
        title_label = QLabel(section_title)
        title_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        header_row.addWidget(title_label)
        header_row.addStretch()

        if toggle_labels and toggle_prefix:
            btn_a = QPushButton(toggle_labels[0])
            btn_b = QPushButton(toggle_labels[1])
            btn_a.setStyleSheet(_BTN_INACTIVE)   # primeiro inativo
            btn_b.setStyleSheet(_BTN_ACTIVE)     # segundo ativo por padrao (receita/total)
            btn_a.setFixedHeight(26)
            btn_b.setFixedHeight(26)
            btn_a.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_b.setCursor(Qt.CursorShape.PointingHandCursor)
            header_row.addWidget(btn_a)
            header_row.addWidget(btn_b)
            setattr(self, f"_{toggle_prefix}_btn_a", btn_a)
            setattr(self, f"_{toggle_prefix}_btn_b", btn_b)

        frame_layout.addLayout(header_row)

        chart_view = self._make_chart_view()
        chart_view.setMinimumHeight(240)
        frame_layout.addWidget(chart_view)

        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: none;
                font-size: 13px;
                color: #2c3e50;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 6px;
                font-weight: bold;
                border: none;
                font-size: 12px;
            }
            QTableWidget::item { color: #2c3e50; padding: 5px 8px; }
        """)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(34)
        table.horizontalHeader().setStretchLastSection(True)
        for i, w in enumerate(col_widths):
            table.setColumnWidth(i, w)

        frame_layout.addWidget(table)

        setattr(self, table_attr, table)
        setattr(self, chart_attr, chart_view)
        return frame

    def _fit_table_height(self, table):
        """Ajusta a altura da tabela para mostrar todas as linhas sem scrollbar."""
        header_h = table.horizontalHeader().height()
        rows_h = sum(table.rowHeight(i) for i in range(table.rowCount()))
        table.setFixedHeight(header_h + rows_h + 4)

    def _update_pie(self, chart_view, labels, values):
        chart = chart_view.chart()
        chart.removeAllSeries()

        total = sum(v for v in values if v > 0)
        if total == 0:
            return

        series = QPieSeries()
        for i, (label, value) in enumerate(zip(labels, values)):
            if value <= 0:
                continue
            pct = value / total * 100
            slc = series.append(f"{label}  {pct:.1f}%", value)
            slc.setColor(QColor(_COLORS[i % len(_COLORS)]))
            slc.setBorderColor(QColor("#ffffff"))
            slc.setBorderWidth(2)

        series.setHoleSize(0.32)   # estilo donut
        chart.addSeries(series)

    # ── DATA ─────────────────────────────────────────────────────────

    def load_report(self):
        self._populate_date_combo()
        self._load_report_for_date(self.current_date)
        self._populate_month_combo()

    def _populate_date_combo(self):
        self.date_combo.blockSignals(True)
        self.date_combo.clear()

        available_dates = self.controller.get_available_dates()

        today = date.today()
        if today not in available_dates:
            available_dates.insert(0, today)

        for d in available_dates:
            label = d.strftime("%d/%m/%Y")
            if d == today:
                label += " (hoje)"
            self.date_combo.addItem(label, d)

        for i in range(self.date_combo.count()):
            if self.date_combo.itemData(i) == today:
                self.date_combo.setCurrentIndex(i)
                break

        self.date_combo.blockSignals(False)

    def _populate_month_combo(self):
        self.month_combo.blockSignals(True)
        self.month_combo.clear()

        months = self.controller.get_available_months()

        from datetime import date as _date
        today = _date.today()
        current = (today.year, today.month)
        if current not in months:
            months.insert(0, current)

        for year, month in months:
            label = f"{_MONTHS_PT[month]} {year}"
            self.month_combo.addItem(label, (year, month))

        self.month_combo.blockSignals(False)

        if self.month_combo.count() > 0:
            ym = self.month_combo.itemData(0)
            self._load_monthly_tops(*ym)

    def _on_month_changed(self, index):
        ym = self.month_combo.itemData(index)
        if ym:
            self._load_monthly_tops(*ym)

    def _load_monthly_tops(self, year, month):
        self._monthly_data = self.controller.get_monthly_tops(year, month)
        self._render_products_panel()
        self._render_customers_panel()

    def _set_sort(self, panel, key):
        if panel == 'products':
            self._products_sort = key
            active_b = (key == 'revenue')
            self._prod_btn_a.setStyleSheet(_BTN_ACTIVE if not active_b else _BTN_INACTIVE)
            self._prod_btn_b.setStyleSheet(_BTN_ACTIVE if active_b else _BTN_INACTIVE)
            self._render_products_panel()
        else:
            self._customers_sort = key
            active_b = (key == 'total_spent')
            self._cust_btn_a.setStyleSheet(_BTN_ACTIVE if not active_b else _BTN_INACTIVE)
            self._cust_btn_b.setStyleSheet(_BTN_ACTIVE if active_b else _BTN_INACTIVE)
            self._render_customers_panel()

    def _render_products_panel(self):
        products = sorted(
            self._monthly_data.get('top_products', []),
            key=lambda x: x[self._products_sort],
            reverse=True,
        )[:5]

        self.top_products_table.setRowCount(len(products))
        for row, p in enumerate(products):
            self.top_products_table.setItem(row, 0, QTableWidgetItem(f"{row + 1}. {p['product_name']}"))
            qty = QTableWidgetItem(str(p['quantity']))
            qty.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.top_products_table.setItem(row, 1, qty)
            rev = QTableWidgetItem(f"R$ {p['revenue']:.2f}")
            rev.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.top_products_table.setItem(row, 2, rev)

        self._update_pie(
            self.top_products_chart_view,
            [p['product_name'] for p in products],
            [p[self._products_sort] for p in products],
        )
        self._fit_table_height(self.top_products_table)

    def _render_customers_panel(self):
        customers = sorted(
            self._monthly_data.get('top_customers', []),
            key=lambda x: x[self._customers_sort],
            reverse=True,
        )[:5]

        self.top_customers_table.setRowCount(len(customers))
        for row, c in enumerate(customers):
            self.top_customers_table.setItem(row, 0, QTableWidgetItem(f"{row + 1}. {c['customer_name']}"))
            orders_item = QTableWidgetItem(str(c['total_orders']))
            orders_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.top_customers_table.setItem(row, 1, orders_item)
            spent = QTableWidgetItem(f"R$ {c['total_spent']:.2f}")
            spent.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.top_customers_table.setItem(row, 2, spent)

        self._update_pie(
            self.top_customers_chart_view,
            [c['customer_name'] for c in customers],
            [c[self._customers_sort] for c in customers],
        )
        self._fit_table_height(self.top_customers_table)

    def on_date_changed(self, index):
        selected_date = self.date_combo.itemData(index)
        if selected_date:
            self.current_date = selected_date
            self._load_report_for_date(selected_date)

    def _load_report_for_date(self, report_date):
        data = self.controller.get_daily_report(report_date)

        # ── Cards ──
        self.card_orders.findChild(QLabel, "value_label").setText(str(data['total_orders']))
        self.card_revenue.findChild(QLabel, "value_label").setText(f"R$ {data['total_revenue']:.2f}")

        total_items = sum(item['quantity'] for item in data['by_product']) if data['by_product'] else 0
        self.card_items.findChild(QLabel, "value_label").setText(str(total_items))

        # ── Categoria ──
        categories = data['by_category']
        self.category_table.setRowCount(len(categories))
        cat_labels, cat_values = [], []
        for row, (cat, info) in enumerate(categories.items()):
            self.category_table.setItem(row, 0, QTableWidgetItem(cat.capitalize()))

            qty = QTableWidgetItem(str(info['quantity']))
            qty.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.category_table.setItem(row, 1, qty)

            rev = QTableWidgetItem(f"R$ {info['revenue']:.2f}")
            rev.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.category_table.setItem(row, 2, rev)

            cat_labels.append(cat.capitalize())
            cat_values.append(info['revenue'])

        self._update_pie(self.category_chart_view, cat_labels, cat_values)

        # ── Produto ──
        products = data['by_product']
        self.product_table.setRowCount(len(products))
        prod_labels, prod_values = [], []
        for row, p in enumerate(products):
            self.product_table.setItem(row, 0, QTableWidgetItem(p['product_name']))

            cat = QTableWidgetItem(p['category'].capitalize())
            cat.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.product_table.setItem(row, 1, cat)

            qty = QTableWidgetItem(str(p['quantity']))
            qty.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.product_table.setItem(row, 2, qty)

            rev = QTableWidgetItem(f"R$ {p['revenue']:.2f}")
            rev.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.product_table.setItem(row, 3, rev)

            prod_labels.append(p['product_name'])
            prod_values.append(p['revenue'])

        self._update_pie(self.product_chart_view, prod_labels, prod_values)

        # ── Comandas ──
        orders = data['order_list']
        self.orders_table.setRowCount(len(orders))
        order_labels, order_values = [], []
        for row, order in enumerate(orders):
            id_item = QTableWidgetItem(str(order['id']))
            id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.orders_table.setItem(row, 0, id_item)

            self.orders_table.setItem(row, 1, QTableWidgetItem(order['customer_name']))

            time_item = QTableWidgetItem(order['created_at'])
            time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.orders_table.setItem(row, 2, time_item)

            total_item = QTableWidgetItem(f"R$ {order['total']:.2f}")
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.orders_table.setItem(row, 3, total_item)

            order_labels.append(order['customer_name'])
            order_values.append(order['total'])

        self._update_pie(self.orders_chart_view, order_labels, order_values)

    # ── EXPORT ───────────────────────────────────────────────────────

    def export_excel(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Erro", "openpyxl nao instalado.\nExecute: pip install openpyxl")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatorio",
            f"relatorio_{self.current_date.strftime('%Y-%m-%d')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not filepath:
            return

        data = self.controller.get_daily_report(self.current_date)

        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Resumo"
        ws.append([f"Relatorio do Dia - {self.current_date.strftime('%d/%m/%Y')}"])
        ws.append([])
        ws.append(["Comandas Fechadas", data['total_orders']])
        ws.append(["Receita Total", data['total_revenue']])
        total_items = sum(item['quantity'] for item in data['by_product']) if data['by_product'] else 0
        ws.append(["Itens Vendidos", total_items])

        ws_cat = wb.create_sheet("Por Categoria")
        ws_cat.append(["Categoria", "Quantidade", "Receita"])
        for cat, info in data['by_category'].items():
            ws_cat.append([cat.capitalize(), info['quantity'], info['revenue']])

        ws_prod = wb.create_sheet("Por Produto")
        ws_prod.append(["Produto", "Categoria", "Quantidade", "Receita"])
        for p in data['by_product']:
            ws_prod.append([p['product_name'], p['category'].capitalize(), p['quantity'], p['revenue']])

        ws_orders = wb.create_sheet("Comandas")
        ws_orders.append(["ID", "Cliente", "Horario", "Total"])
        for order in data['order_list']:
            ws_orders.append([order['id'], order['customer_name'], order['created_at'], order['total']])

        wb.save(filepath)
        QMessageBox.information(self, "Sucesso", f"Relatorio exportado:\n{filepath}")
